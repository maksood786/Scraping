from selenium import webdriver
import openpyxl
from selenium.webdriver.common.keys import Keys
import time
import pandas as pd
import os
from random import randint
import datetime

		
if __name__ == '__main__':

        now=datetime.datetime.now()
        filename='Output'+'_'+str(now)+'.xlsx'
        filename=filename.replace(":",'-')
        wb = openpyxl.Workbook()  
        sheet = wb['Sheet']  
        sheet.cell(row=1, column=1).value='Link'
        sheet.cell(row=1, column=2).value='title'
        sheet.cell(row=1, column=3).value='address'
        sheet.cell(row=1, column=4).value='Bedrooms'
        sheet.cell(row=1, column=5).value='Residences'
        sheet.cell(row=1, column=6).value='Estimated completion'
        sheet.cell(row=1, column=7).value='Buildings'
        sheet.cell(row=1, column=8).value='Architect'
        sheet.cell(row=1, column=9).value='Builder'
        sheet.cell(row=1, column=10).value='Local Government Area'
        sheet.cell(row=1, column=11).value='Suburb'
        sheet.cell(row=1, column=12).value='Project Status'
        sheet.cell(row=1, column=13).value='Listing Type'
        sheet.cell(row=1, column=14).value='Building Type'
        sheet.cell(row=1, column=15).value='Estimated Completion Date'
        sheet.cell(row=1, column=16).value='Nearby Public Transport'
        sheet.cell(row=1, column=17).value='Floor Count'
        sheet.cell(row=1, column=18).value='Number of Buildings'
        sheet.cell(row=1, column=19).value='Number of Residences'
        sheet.cell(row=1, column=20).value='Developer Name'
        sheet.cell(row=1, column=21).value='Developer info'
        sheet.cell(row=1, column=22).value='Architect Name'
        sheet.cell(row=1, column=23).value='Architect Info'
        sheet.cell(row=1, column=24).value='Builder Name'
        sheet.cell(row=1, column=25).value='Builder Info'
        sheet.cell(row=1, column=26).value='Building Type'
        sheet.cell(row=1, column=27).value='Floor Count'

        Tickers=pd.read_excel('file:///' + os.path.join(os.path.dirname(__file__)) + '/Links.xlsx',sheet_name='Link')    
        
        
        driver = webdriver.Chrome()       
        
        r=2
        for Tickers_Loop in range(len(Tickers)):                
            Link=Tickers.loc[Tickers_Loop,'Link']
            print(Link)
            driver.get(Link)
            
            time.sleep(5)
            
            body = driver.find_element_by_css_selector('body')
            
            body.send_keys(Keys.PAGE_DOWN)
            
            time.sleep(5)
            
            try:            
                driver.find_element_by_xpath('//*[@id="project-development-info"]/div/div/div/div/div/div[2]/a').click()
            except:
                pass

            try:
                title = driver.find_element_by_class_name('project-title__text').text

            except:
                title=''
                
            try:
                address = driver.find_element_by_class_name('project-address').text                
            except:
                address=''

            try:
                Type1 = driver.find_element_by_class_name('Building type').text                
            except:
                Type1=''

            try:
                FloorCount = driver.find_element_by_class_name('Floor count').text                
            except:
                FloorCount=''
                
            
            try:
                elementl = driver.find_element_by_xpath("//*[text()='Bedrooms']")
                element2=elementl.find_element_by_xpath('..')
                data1=element2.find_element_by_tag_name('p').text
            except:
                data1=''

            try:
                elementl = driver.find_element_by_xpath("//*[text()='Residences']")
                element2=elementl.find_element_by_xpath('..')
                data2=element2.find_element_by_tag_name('p').text
            except:
                data2=''    
                
            try:
                elementl = driver.find_element_by_xpath("//*[text()='Estimated completion']")
                element2=elementl.find_element_by_xpath('..')
                data3=element2.find_element_by_tag_name('p').text
            except:
                data3=''


            try:
                elementl = driver.find_element_by_xpath("//*[text()='Buildings']")
                element2=elementl.find_element_by_xpath('..')
                data4=element2.find_element_by_tag_name('p').text
            except:
                data4=''      
                
            
            

            try:
                elementl = driver.find_element_by_xpath("//*[text()=' Architect:']")
                element2=elementl.find_element_by_xpath('..')
                data5=element2.find_element_by_class_name('value').text                
            except:                
                data5=''     

            try:
                elementl = driver.find_element_by_xpath("//*[text()=' Builder:']")
                element2=elementl.find_element_by_xpath('..')
                data6=element2.find_element_by_class_name('value').text                
            except:                
                data6='' 

            try:
                elementl = driver.find_element_by_xpath("//*[text()=' Local Government Area:']")
                element2=elementl.find_element_by_xpath('..')
                data7=element2.find_element_by_class_name('value').text                
            except:                
                data7='' 

            try:
                elementl = driver.find_element_by_xpath("//*[text()=' Suburb:']")
                element2=elementl.find_element_by_xpath('..')
                data8=element2.find_element_by_class_name('value').text                
            except:                
                data8=''                 

            try:
                elementl = driver.find_element_by_xpath("//*[text()=' Project Status:']")
                element2=elementl.find_element_by_xpath('..')
                data9=element2.find_element_by_class_name('value').text                
            except:                
                data9='' 

            try:
                elementl = driver.find_element_by_xpath("//*[text()=' Listing Type:']")
                element2=elementl.find_element_by_xpath('..')
                data10=element2.find_element_by_class_name('value').text                
            except:                
                data10='' 

                
            try:
                elementl = driver.find_element_by_xpath("//*[text()=' Building Type:']")
                element2=elementl.find_element_by_xpath('..')
                data11=element2.find_element_by_class_name('value').text  
                print(data11)
            except:                
                data11='' 

            try:
                elementl = driver.find_element_by_xpath("//*[text()=' Estimated Completion Date:']")
                element2=elementl.find_element_by_xpath('..')
                data12=element2.find_element_by_class_name('value').text                
            except:       
                print('error')
                data12='' 
                
     

            try:

                elementl = driver.find_element_by_xpath("//*[contains(text(), 'Nearby Public Transport:')]")
                
                element2=elementl.find_element_by_xpath('..')
                data13=element2.find_element_by_class_name('value').text                
            except:
               
                data13='' 
            

            try:
                elementl = driver.find_element_by_xpath("//*[text()=' Floor Count:']")
                element2=elementl.find_element_by_xpath('..')
                data14=element2.find_element_by_class_name('value').text                
            except:                
                data14='' 

            try:
                elementl = driver.find_element_by_xpath("//*[text()=' Number of Buildings:']")
                element2=elementl.find_element_by_xpath('..')
                data15=element2.find_element_by_class_name('value').text                
            except:                
                data15=''                 

            try:
                elementl = driver.find_element_by_xpath("//*[text()=' Number of Residences:']")
                element2=elementl.find_element_by_xpath('..')
                data16=element2.find_element_by_class_name('value').text                
            except:                
                data16=''      
                
            try:
                data17=driver.find_elements_by_class_name('developer-name')[0].text               
                data18=driver.find_elements_by_class_name('developer-detail-info-text')[0].text          
            except:
                pass                
#                data17=''
#                data18=''
            
            try:
                driver.find_element_by_xpath('//*[@id="ui-id-2"]').click()
                time.sleep(3)
            except:
                pass
          

            try:
                data19=driver.find_elements_by_class_name('developer-name')[1].text
                data20=driver.find_elements_by_class_name('developer-detail-info-text')[1].text          
            except:
                pass
#                data19=''
#                data20=''

                
            try:
                driver.find_element_by_xpath('//*[@id="ui-id-3"]').click()
                time.sleep(3)
            except:
                pass
            
            
            try:
                data21=driver.find_elements_by_class_name('developer-name')[2].text
                data22=driver.find_elements_by_class_name('developer-detail-info-text')[2].text          
            except:
                pass
#                print('error21')                
#                data21=''
#                data22=''
            
            sheet.cell(row=r, column=1).value=Link
            sheet.cell(row=r, column=2).value=title
            sheet.cell(row=r, column=3).value=address
            sheet.cell(row=r, column=4).value=data1
            sheet.cell(row=r, column=5).value=data2
            sheet.cell(row=r, column=6).value=data3
            sheet.cell(row=r, column=7).value=data4
            sheet.cell(row=r, column=8).value=data5
            sheet.cell(row=r, column=9).value=data6
            sheet.cell(row=r, column=10).value=data7
            sheet.cell(row=r, column=11).value=data8
            sheet.cell(row=r, column=12).value=data9
            sheet.cell(row=r, column=13).value=data10
            sheet.cell(row=r, column=14).value=data11
            sheet.cell(row=r, column=15).value=data12
            sheet.cell(row=r, column=16).value=data13
            sheet.cell(row=r, column=17).value=data14
            sheet.cell(row=r, column=18).value=data15
            sheet.cell(row=r, column=19).value=data16
            sheet.cell(row=r, column=20).value=data17
            sheet.cell(row=r, column=21).value=data18
            sheet.cell(row=r, column=22).value=data19
            sheet.cell(row=r, column=23).value=data20
            sheet.cell(row=r, column=24).value=data21
            sheet.cell(row=r, column=25).value=data22
            sheet.cell(row=r, column=26).value=Type1
            sheet.cell(row=r, column=27).value=FloorCount
            
            wb.save(filename)
            
            
            r=r+1
        
        
        driver.quit()
        
        
       

        
